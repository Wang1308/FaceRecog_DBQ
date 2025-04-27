from flask import Flask, render_template, request, redirect, send_file,flash
from database import connect_db
from datetime import datetime, date
import os
import openpyxl
from flask_cors import CORS

app = Flask(__name__)
CORS(app)
app.secret_key = 'bi_mat_cua_wang'

@app.route('/', methods=['GET', 'POST'])
def index():
    selected_date = request.form.get('date') if request.method == 'POST' else date.today().strftime('%Y-%m-%d')
    conn = connect_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT e.id, e.name, a.checkin_time 
        FROM attendance a 
        JOIN employees e ON e.id = a.employee_id 
        WHERE a.checkin_time::date = %s
    """, (selected_date,))
    today_records = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('index.html', records=today_records, selected_date=selected_date)

@app.route('/add', methods=['GET', 'POST'])
def add_employee():
    if request.method == 'POST':
        name = request.form['name']
        department = request.form['department']
        position = request.form['position']

        # Nhận toàn bộ các file ảnh trong folder
        files = request.files.getlist('images')

        # Kiểm tra xem có đủ 5 ảnh không
        if len(files) != 5:
            return "Vui lòng upload đúng 5 ảnh trong folder!"

        # Đọc từng file ảnh
        images = [f.read() for f in files]

        # Kết nối cơ sở dữ liệu
        conn = connect_db()
        cur = conn.cursor()

        # Kiểm tra xem nhân viên đã tồn tại hay chưa
        cur.execute("SELECT * FROM employees WHERE name=%s AND department=%s AND position=%s", (name, department, position))
        if cur.fetchone():
            return "Nhân viên đã tồn tại!"

        # Chèn thông tin vào bảng employees
        cur.execute("""
            INSERT INTO employees (name, department, position, face_image1, face_image2, face_image3, face_image4, face_image5)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (name, department, position, *images))

        # Lưu thay đổi vào cơ sở dữ liệu
        conn.commit()
        cur.close()
        conn.close()

        return redirect('/add?success=1')

    return render_template('add_employee.html')


@app.route('/employees')
def list_employees():
    conn = connect_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name, department, position FROM employees")
    employees = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('list_employees.html', employees=employees)

@app.route('/delete', methods=['GET', 'POST'])
def delete_employee():
    conn = connect_db()
    cur = conn.cursor()

    # Tìm kiếm nhân viên theo ID
    if request.method == 'POST' and 'search_id' in request.form:
        search_id = request.form['search_id']
        cur.execute("SELECT id, name, department, position FROM employees WHERE id = %s", (search_id,))
        employee = cur.fetchone()

        # Kiểm tra xem có tìm thấy nhân viên không
        if employee:
            return render_template('delete_employee.html', employee=employee, searched=True)
        else:
            flash("Không tìm thấy nhân viên với ID này.", 'danger')
            return render_template('delete_employee.html', searched=True)

    # Xử lý khi nhấn nút xóa
    if request.method == 'POST' and 'delete_id' in request.form:
        delete_id = request.form['delete_id']
        # Xóa nhân viên khỏi cơ sở dữ liệu
        cur.execute("DELETE FROM employees WHERE id = %s", (delete_id,))
        conn.commit()
        flash("Nhân viên đã được xóa thành công.", 'success')
        return redirect('/delete')  # Điều hướng lại để xem kết quả sau khi xóa

    cur.close()
    conn.close()
    return render_template('delete_employee.html')

@app.route('/monthly', methods=['GET', 'POST'])
def monthly_report():
    selected_month = request.form.get('month') if request.method == 'POST' else date.today().strftime('%Y-%m')
    conn = connect_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT e.id, e.name, a.checkin_time 
        FROM attendance a 
        JOIN employees e ON e.id = a.employee_id 
        WHERE TO_CHAR(a.checkin_time, 'YYYY-MM') = %s
    """, (selected_month,))
    records = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('monthly_report.html', records=records, selected_month=selected_month)

@app.route('/export')
def export_excel():
    selected_month = request.args.get('month')
    if not selected_month:
        selected_month = date.today().strftime('%Y-%m')

    conn = connect_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT e.id, e.name, 
               DATE(a.checkin_time) AS checkin_date, 
               TO_CHAR(a.checkin_time, 'HH24:MI:SS') AS checkin_time
        FROM attendance a 
        JOIN employees e ON e.id = a.employee_id
        WHERE TO_CHAR(a.checkin_time, 'YYYY-MM') = %s
    """, (selected_month,))
    data = cur.fetchall()
    cur.close()
    conn.close()

    # Tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['ID', 'Tên', 'Ngày chấm công', 'Giờ chấm công'])
    for row in data:
        ws.append(row)

    for col in range(1, 5):
        max_length = 0
        column = openpyxl.utils.get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        ws.column_dimensions[column].width = max_length + 2

    path = "attendance_report.xlsx"
    wb.save(path)
    return send_file(path, as_attachment=True)



if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port = '5000')
